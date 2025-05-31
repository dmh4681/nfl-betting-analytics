"""
NFL 2025 Season Setup Script
Creates the Excel structure and initial data for the 2025 NFL season
"""

import pandas as pd
import json
import os
from pathlib import Path
from datetime import datetime, timedelta
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class NFL2025SeasonSetup:
    def __init__(self):
        """Initialize the season setup"""
        # Get the actual project root - go up from scripts/ directory
        self.project_root = Path(__file__).parent.parent
        self.config_path = self.project_root / "config" / "team_mappings.json"
        self.output_path = self.project_root / "data" / "current_season" / "NFL_2025_Season.xlsx"
        
        logger.info(f"Project root: {self.project_root}")
        logger.info(f"Config path: {self.config_path}")
        logger.info(f"Output path: {self.output_path}")
        
        # Load team mappings
        self.team_mappings = self._load_team_mappings()
        
        # 2025 NFL Schedule (key dates)
        self.season_start = datetime(2025, 9, 4)  # Estimated - typically first Thursday in September
        self.regular_season_weeks = 18
        self.playoff_weeks = ["Wild Card", "Divisional", "Conference Championship", "Super Bowl"]
        
    def _load_team_mappings(self):
        """Load team mappings from config"""
        try:
            with open(self.config_path, 'r') as f:
                mappings = json.load(f)
                logger.info(f"Loaded {len(mappings)} team mappings")
                return mappings
        except FileNotFoundError:
            logger.error(f"Team mappings not found at {self.config_path}")
            return {}
            
    def create_weekly_game_template(self, week_num=None, week_name=None):
        """
        Create template for a weekly games sheet
        
        Args:
            week_num: Week number (1-18 for regular season, None for playoffs)
            week_name: Optional name override (e.g., "Wild Card")
        """
        if week_name is None and week_num is not None:
            week_name = f"Week {week_num}"
        elif week_name is None:
            week_name = "Game Week"
            
        # Standard columns for game tracking
        columns = [
            'Away Tag', 'Home Tag', 'Away Team', 'Home Team',
            'Projected Line (Power Rankings)', '"Edge" Over Current Line',
            'Look Ahead Home Line', 'Opening Line', 'Current Line', 'Closing Line',
            'Look Ahead Total', 'Opening Total', 'Current Total', 'Closing Total',
            'Justin Comments', 'Dylan Comments',
            # Betting tracking columns
            'Dylan', 'Type', 'Game', 'Wager', 'American Odds', 'To Decimal Odds',
            'Imp. Probability', '# units bet', 'To win', 'win/lose'
        ]
        
        # Determine number of games
        if week_num is not None and week_num <= 18:
            num_games = 16  # Regular season
        else:
            # Playoff games - vary by round
            playoff_game_counts = {
                "Wild Card": 6,
                "Divisional": 4, 
                "Conference Championship": 2,
                "Super Bowl": 1
            }
            num_games = playoff_game_counts.get(week_name, 8)
        
        # Create DataFrame with empty rows for games
        df = pd.DataFrame(columns=columns)
        
        # Add some standard rows
        for i in range(num_games):
            df.loc[i] = [None] * len(columns)
            
        return df, week_name
        
    def create_betting_log_template(self, player="Dylan"):
        """Create template for betting performance tracking"""
        
        columns = [
            'Week', 'Game(s)', 'Wager', 'Book(s)', 'Type', 'US Odds', 'Price', 'Stake',
            'iProb', 'Result', 'Units Won', 'Running Total', 'CL', 'Imp Prob',
            'OCL', 'Imp Prob', 'Total', 'NV iProb', 'Difference', 'nvCLV', 'phony CLV',
            'Notes / Calculations'
        ]
        
        df = pd.DataFrame(columns=columns)
        
        # Add preseason placeholder
        df.loc[0] = ['Preseason', None, None, None, None, None, 1, None, 1, None, 
                     '#N/A', '#N/A', None, 0, None, 0, 0, 0, -1, -1, -1, None]
        
        return df
        
    def create_power_rankings_template(self):
        """Create template for power rankings tracking"""
        
        columns = [
            'Index', 'Team Abbrev', 'Team Name', 'Team CAPS', 
            '2025 Current Rating', 'Rank 2025PF', '2025 NFL Power Rank Rating', 'Rank 2025PR',
            'var PF to PR', "Imp. '24 Record", '2024 End Season Rating', 'Rank 2024',
            'var PF to 2024', 'var PR to 2024'
        ]
        
        df = pd.DataFrame(columns=columns)
        
        # Add all 32 teams
        for i, (team_key, team_data) in enumerate(self.team_mappings.items(), 1):
            df.loc[i-1] = [
                i, team_data['abbreviation'], team_data['full_name'], team_data['caps'],
                None, None, None, None, None, None, None, None, None, None
            ]
            
        return df
        
    def create_season_excel_file(self):
        """Create the complete 2025 season Excel file"""
        
        logger.info("Creating 2025 NFL Season Excel file...")
        
        # Create new workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # 1. Create regular season weeks (1-18)
        for week in range(1, self.regular_season_weeks + 1):
            df, sheet_name = self.create_weekly_game_template(week_num=week)
            ws = wb.create_sheet(f"2025 Week {week}")
            
            # Add dataframe to sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
                
            logger.info(f"Created Week {week}")
            
        # 2. Create playoff weeks
        for playoff_week in self.playoff_weeks:
            df, sheet_name = self.create_weekly_game_template(week_name=playoff_week)
            ws = wb.create_sheet(f"2025 {playoff_week}")
            
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
                
            logger.info(f"Created {playoff_week}")
            
        # 3. Create betting logs
        for player in ["Dylan", "Justin"]:
            df = self.create_betting_log_template(player)
            ws = wb.create_sheet(f"NFL 2025 - {player}")
            
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
                
            logger.info(f"Created betting log for {player}")
            
        # 4. Create power rankings
        df = self.create_power_rankings_template()
        ws = wb.create_sheet("2025 Power Rankings")
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
            
        logger.info("Created power rankings sheet")
        
        # 5. Create schedule placeholder
        ws = wb.create_sheet("2025 Schedule")
        ws.append(["Week", "Date", "Time (ET)", "Away Team", "Home Team", 
                  "Away Abbrev", "Home Abbrev", "TV", "Stadium"])
        logger.info("Created schedule placeholder")
        
        # Ensure output directory exists
        os.makedirs(self.output_path.parent, exist_ok=True)
        
        # Save workbook
        wb.save(self.output_path)
        logger.info(f"2025 NFL Season file created: {self.output_path}")
        
        return self.output_path
        
    def create_season_summary(self):
        """Create a summary of what was set up"""
        
        summary = {
            "season": 2025,
            "created_date": datetime.now().isoformat(),
            "sheets_created": {
                "regular_season_weeks": self.regular_season_weeks,
                "playoff_weeks": len(self.playoff_weeks),
                "betting_logs": 2,  # Dylan and Justin
                "power_rankings": 1,
                "schedule": 1
            },
            "total_sheets": self.regular_season_weeks + len(self.playoff_weeks) + 4,
            "teams_mapped": len(self.team_mappings),
            "file_location": str(self.output_path)
        }
        
        # Save summary
        summary_path = self.output_path.parent / "2025_season_summary.json"
        with open(summary_path, 'w') as f:
            json.dump(summary, f, indent=2)
            
        logger.info(f"Season summary saved: {summary_path}")
        return summary


def main():
    """Set up the 2025 NFL season"""
    
    print("🏈 Setting up 2025 NFL Season...")
    
    # Initialize setup
    setup = NFL2025SeasonSetup()
    
    # Create the Excel file
    excel_path = setup.create_season_excel_file()
    
    # Create summary
    summary = setup.create_season_summary()
    
    print(f"\n✅ 2025 NFL Season Setup Complete!")
    print(f"📁 File created: {excel_path}")
    print(f"📊 Total sheets: {summary['total_sheets']}")
    print(f"🏈 Teams mapped: {summary['teams_mapped']}")
    print(f"📅 Regular season weeks: {summary['sheets_created']['regular_season_weeks']}")
    print(f"🏆 Playoff weeks: {summary['sheets_created']['playoff_weeks']}")
    
    print(f"\n📋 Next steps:")
    print(f"1. Add actual 2025 NFL schedule when released")
    print(f"2. Set up automated line collection")
    print(f"3. Configure CLV calculations")
    
    return excel_path


if __name__ == "__main__":
    main()
